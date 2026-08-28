"""
Batch evaluation from a spreadsheet.

A reviewer does not want to run twelve commands. They want a sheet with one row
per screen, and they want to hand that sheet to someone else afterwards. So this
module takes a workbook with the columns a reviewer already keeps, fills in what
Brush found, and scores each row on a 0 / 0.5 / 1 rubric that is written into the
sheet as a live formula rather than baked in as a number.

Two different scores live here and conflating them is the mistake to avoid:

  CONFORMANCE   Does the code match the design? Brush's verdict on the artefact.
                1.0  no blockers, no majors — ships as designed
                0.5  no blockers, at least one major — partially conforms
                0.0  at least one blocker — accessibility or a broken state

  POINTS        Was Brush right? Only computed when the sheet carries an EXPECTED
                column filled in by a human. Exact agreement scores 1.0, half a
                band apart scores 0.5, a full band apart scores 0.0.

Without EXPECTED, POINTS mirrors CONFORMANCE and the sheet says so, because a
model that grades its own homework and reports the mark as an accuracy figure is
the exact failure this project exists to argue against.
"""
from __future__ import annotations

import os
import time
from dataclasses import dataclass, field
from typing import Any, Optional

from ..agents.orchestrator import run_audit
from ..extract.vision import resolve_design_ref

FONT = "Arial"

# The column names a reviewer keeps, and the aliases we accept for each so an
# existing sheet does not have to be renamed to be usable.
COLUMNS = {
    "id": ["id", "case", "case id", "caso"],
    "design": ["image/figma", "image", "figma", "design", "mockup", "diseño", "imagen"],
    "code": ["code", "codigo", "código", "implementation", "html"],
    "css": ["css", "stylesheet", "styles"],
    "response": ["respuesta brush", "brush response", "respuesta", "response", "output"],
    "conformance": ["conformance", "conformidad"],
    "points": ["points", "puntos", "score", "puntaje"],
    "expected": ["expected", "esperado", "human", "ground truth", "gt"],
    "comments": ["comments", "comentarios", "notes", "notas"],
}

HEADERS = ["ID", "IMAGE/FIGMA", "CODE", "CSS", "BRUSH RESPONSE",
           "BLOCKERS", "MAJORS", "MINORS", "INFO",
           "CONFORMANCE", "EXPECTED", "POINTS", "COMMENTS"]

INK = "1F2430"
RULE = "D9DDE3"
HEAD_FILL = "1F2430"
INPUT_FILL = "FFF9DB"     # cells a human fills in
OUT_FILL = "F4F6F8"       # cells Brush writes


class SheetFormatError(Exception):
    """The workbook cannot be read; the person can fix the header row."""


class MissingDependencyError(Exception):
    """Spreadsheet mode was reached without openpyxl installed."""


@dataclass
class JobResult:
    case_id: str
    design_ref: str
    code_ref: str
    css_refs: list[str]
    ok: bool
    response: str = ""
    comments: str = ""
    counts: dict[str, int] = field(default_factory=dict)
    conformance: Optional[float] = None
    findings: list = field(default_factory=list)
    run_id: str = ""
    wall_seconds: float = 0.0
    cost_usd: float = 0.0
    note: str = ""
    error: str = ""


# ---------------------------------------------------------------------------
# Reading
# ---------------------------------------------------------------------------
def _openpyxl():
    """
    Import openpyxl only when a workbook is actually touched.

    The scoring rules in this module -- `parse_expected`, `agreement_points`,
    `conformance_score` -- are pure arithmetic with no spreadsheet in sight, and
    they are exercised by the CLI test suite. A module-level import meant those
    tests crashed on any machine without the optional dependency installed.
    """
    try:
        from openpyxl import Workbook, load_workbook
        from openpyxl.comments import Comment
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    except ImportError as exc:
        raise MissingDependencyError(
            "spreadsheet mode needs openpyxl, which is not installed"
        ) from exc
    return Workbook, load_workbook, Comment, Alignment, Border, Font, PatternFill, Side


def _norm(s: Any) -> str:
    return str(s or "").strip().lower()


def _map_headers(row: list) -> dict[str, int]:
    found: dict[str, int] = {}
    for idx, cell in enumerate(row):
        label = _norm(cell)
        if not label:
            continue
        for key, aliases in COLUMNS.items():
            if label in aliases and key not in found:
                found[key] = idx
    return found


def read_jobs(path: str) -> tuple[list[dict], dict[str, int]]:
    _, load_workbook, *_ = _openpyxl()
    wb = load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise SheetFormatError("the first sheet of the workbook is empty")

    header_idx = 0
    cols: dict[str, int] = {}
    for i, row in enumerate(rows[:10]):
        candidate = _map_headers(list(row))
        if "id" in candidate and ("design" in candidate or "code" in candidate):
            header_idx, cols = i, candidate
            break
    if not cols:
        seen = [str(c) for c in rows[0][:8] if c] if rows else []
        raise SheetFormatError(
            "could not find the header row in the first sheet. The first row needs "
            "a column named ID, plus IMAGE/FIGMA or CODE."
            + (f" Found instead: {', '.join(seen)}." if seen else "")
        )

    jobs = []
    for row in rows[header_idx + 1:]:
        def get(key: str) -> Any:
            i = cols.get(key)
            return row[i] if i is not None and i < len(row) else None

        if not any(get(k) for k in ("id", "design", "code")):
            continue
        jobs.append({
            "id": str(get("id") or f"row-{len(jobs) + 1}").strip(),
            "design": str(get("design") or "").strip(),
            "code": str(get("code") or "").strip(),
            "css": str(get("css") or "").strip(),
            "expected": get("expected"),
        })
    return jobs, cols


# ---------------------------------------------------------------------------
# Scoring
# ---------------------------------------------------------------------------
def conformance_score(counts: dict[str, int]) -> float:
    if counts.get("blocker", 0) > 0:
        return 0.0
    if counts.get("major", 0) > 0:
        return 0.5
    return 1.0


def parse_expected(value: Any) -> tuple[Optional[float], str]:
    """
    Coerce whatever a human typed into a score, or explain why it is not one.

    A text value left in the EXPECTED column produced `#VALUE!` in the delivered
    workbook, because the POINTS formula does arithmetic on it. Accept the
    handful of spellings people actually use, and keep anything else out of the
    cell entirely.
    """
    if value in (None, ""):
        return None, ""
    if isinstance(value, bool):
        return (1.0 if value else 0.0), ""
    if isinstance(value, (int, float)):
        v = float(value)
        return (v, "") if v in (0.0, 0.5, 1.0) else (
            None, f"EXPECTED must be 0, 0.5 or 1 — got {v:g}")
    text = str(value).strip().lower().replace(",", ".")
    words = {
        "1": 1.0, "1.0": 1.0, "ok": 1.0, "pass": 1.0, "yes": 1.0, "si": 1.0,
        "sí": 1.0, "conforms": 1.0, "cumple": 1.0,
        "0.5": 0.5, ".5": 0.5, "partial": 0.5, "parcial": 0.5,
        "0": 0.0, "0.0": 0.0, "fail": 0.0, "no": 0.0, "falla": 0.0,
    }
    if text in words:
        return words[text], ""
    return None, f"EXPECTED not understood: {value!r} — use 0, 0.5 or 1"


def agreement_points(conformance: Optional[float], expected: Any) -> Optional[float]:
    """Reward exact agreement, half-credit for one band, nothing for two."""
    exp, _ = parse_expected(expected)
    if conformance is None or exp is None:
        return None
    gap = abs(conformance - exp)
    if gap < 1e-9:
        return 1.0
    if abs(gap - 0.5) < 1e-9:
        return 0.5
    return 0.0


def build_comment(res: JobResult) -> str:
    """
    The COMMENTS cell. Says what drove the score and what to do about it, which
    is the only reason a reviewer reads this column.
    """
    if not res.ok:
        return f"Not audited — {res.error}"
    c = res.counts
    if res.conformance == 1.0:
        base = "Matches the specification on every measured property."
        if c.get("minor") or c.get("info"):
            base += (f" {c.get('minor', 0)} minor and {c.get('info', 0)} informational "
                     f"note(s) recorded; none change what a user sees.")
        return base + (f" {res.note}" if res.note else "")

    parts = []
    blockers = [f for f in res.findings if f.severity == "blocker"]
    majors = [f for f in res.findings if f.severity == "major"]
    if blockers:
        parts.append(f"{len(blockers)} blocker(s): " +
                     "; ".join(f.title for f in blockers[:2]) +
                     ("…" if len(blockers) > 2 else ""))
    if majors:
        parts.append(f"{len(majors)} major deviation(s), led by {majors[0].title}")
    top = (blockers or majors or res.findings)
    if top and top[0].suggested_fix:
        parts.append(f"Start with: {top[0].suggested_fix}")
    if res.note:
        parts.append(res.note)
    return " ".join(parts)


def build_response(res: JobResult) -> str:
    if not res.ok:
        return "ERROR"
    c = res.counts
    total = sum(c.values())
    if total == 0:
        return "Conforms — no deviations found."
    bits = [f"{c[k]} {k}" for k in ("blocker", "major", "minor", "info") if c.get(k)]
    return f"{total} finding(s): " + ", ".join(bits)


# ---------------------------------------------------------------------------
# Running
# ---------------------------------------------------------------------------
def run_job(job: dict, provider, base_dir: str, out_dir: str,
            ledger: Optional[str], accept_drafted: bool) -> JobResult:
    res = JobResult(case_id=job["id"], design_ref=job["design"],
                    code_ref=job["code"], css_refs=[], ok=False)
    t0 = time.time()
    try:
        design_path, note = resolve_design_ref(
            job["design"], provider, accept_drafted, base_dir)
        res.note = note

        code_path = job["code"]
        if code_path and not os.path.isabs(code_path):
            code_path = os.path.join(base_dir, code_path)
        if not code_path or not os.path.exists(code_path):
            raise FileNotFoundError(f"implementation not found: {job['code']}")

        css = [c.strip() for c in (job["css"] or "").replace(";", ",").split(",") if c.strip()]
        css = [c if os.path.isabs(c) else os.path.join(base_dir, c) for c in css]
        if not css:
            guess = os.path.splitext(code_path)[0] + ".css"
            if os.path.exists(guess):
                css = [guess]
        missing = [c for c in css if not os.path.exists(c)]
        if missing:
            raise FileNotFoundError(f"stylesheet not found: {missing[0]}")
        if not css:
            raise FileNotFoundError("no stylesheet given and none found next to the HTML")
        res.css_refs = css

        rep = run_audit(design_path, code_path, css, provider, out_dir=out_dir,
                        ledger_path=ledger, ignore_annotations=True,
                        run_id=f"batch-{job['id']}")
        res.ok = True
        res.findings = rep.findings
        res.counts = {k: rep.stats["by_severity"].get(k, 0)
                      for k in ("blocker", "major", "minor", "info")}
        res.conformance = conformance_score(res.counts)
        res.run_id = rep.run_id
        res.cost_usd = rep.stats["usage"]["cost_usd"]
    except Exception as exc:
        res.error = f"{type(exc).__name__}: {exc}"
    res.wall_seconds = round(time.time() - t0, 3)
    res.response = build_response(res)
    res.comments = build_comment(res)
    return res


# ---------------------------------------------------------------------------
# Writing
# ---------------------------------------------------------------------------
def _style_header(ws, ncols: int) -> None:
    _, _, Comment, Alignment, Border, Font, PatternFill, Side = _openpyxl()
    thin = Side(style="thin", color=RULE)
    for c in range(1, ncols + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = Font(name=FONT, bold=True, size=10, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=HEAD_FILL)
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        cell.border = Border(bottom=thin)
    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "A2"


def write_workbook(results: list[JobResult], out_path: str,
                   jobs: list[dict], provider_name: str) -> str:
    Workbook, _, Comment, Alignment, Border, Font, PatternFill, Side = _openpyxl()
    wb = Workbook()
    ws = wb.active
    ws.title = "Cases"

    ws.append(HEADERS)
    _style_header(ws, len(HEADERS))

    thin = Side(style="thin", color=RULE)
    has_expected = any(parse_expected(j.get("expected"))[0] is not None for j in jobs)

    for i, (res, job) in enumerate(zip(results, jobs), start=2):
        c = res.counts
        ws.cell(row=i, column=1, value=res.case_id)
        ws.cell(row=i, column=2, value=res.design_ref)
        ws.cell(row=i, column=3, value=res.code_ref)
        ws.cell(row=i, column=4, value=", ".join(os.path.basename(x) for x in res.css_refs))
        ws.cell(row=i, column=5, value=res.response)
        ws.cell(row=i, column=6, value=c.get("blocker", 0) if res.ok else None)
        ws.cell(row=i, column=7, value=c.get("major", 0) if res.ok else None)
        ws.cell(row=i, column=8, value=c.get("minor", 0) if res.ok else None)
        ws.cell(row=i, column=9, value=c.get("info", 0) if res.ok else None)

        # CONFORMANCE is a live formula, not a baked number, so a reviewer can
        # see the rubric and the sheet still recalculates if counts are edited.
        ws.cell(row=i, column=10,
                value=f'=IF(F{i}="","",IF(F{i}>0,0,IF(G{i}>0,0.5,1)))')
        exp_value, exp_note = parse_expected(job.get("expected"))
        exp_cell = ws.cell(row=i, column=11, value=exp_value)
        if exp_note:
            exp_cell.comment = Comment(
                f"{exp_note}\nOriginal entry: {job.get('expected')!r}", "Brush")
        ws.cell(row=i, column=12,
                value=(f'=IF(OR(J{i}="",K{i}=""),"",'
                       f'IF(ABS(J{i}-K{i})<0.001,1,IF(ABS(ABS(J{i}-K{i})-0.5)<0.001,0.5,0)))')
                if has_expected else f'=IF(J{i}="","",J{i})')
        ws.cell(row=i, column=13, value=res.comments)

        for col in range(1, len(HEADERS) + 1):
            cell = ws.cell(row=i, column=col)
            cell.font = Font(name=FONT, size=10)
            cell.alignment = Alignment(vertical="top", wrap_text=col in (5, 13))
            cell.border = Border(bottom=thin)
            if col in (5, 6, 7, 8, 9, 10, 13):
                cell.fill = PatternFill("solid", fgColor=OUT_FILL)
            if col == 11:
                cell.fill = PatternFill("solid", fgColor=INPUT_FILL)
            if col in (10, 12):
                cell.number_format = "0.0"
        ws.row_dimensions[i].height = 46

    ws.cell(row=1, column=10).comment = Comment(
        "Does the code match the design?\n"
        "1.0  no blockers, no majors\n"
        "0.5  no blockers, at least one major\n"
        "0.0  at least one blocker\n"
        "Formula, not a stored number — edit the counts and it recalculates.",
        "Brush")
    ws.cell(row=1, column=11).comment = Comment(
        "Fill this in yourself: what you judged the row to be (0, 0.5 or 1).\n"
        "Leave blank if you are not scoring Brush.", "Brush")
    ws.cell(row=1, column=12).comment = Comment(
        "Was Brush right? Only meaningful when EXPECTED is filled in.\n"
        "1.0 exact agreement · 0.5 one band apart · 0.0 two bands apart.\n"
        "With EXPECTED blank this mirrors CONFORMANCE and measures nothing.",
        "Brush")

    for col, width in zip("ABCDEFGHIJKLM",
                          [14, 30, 26, 22, 30, 10, 9, 9, 8, 14, 12, 10, 62]):
        ws.column_dimensions[col].width = width

    _findings_sheet(wb, results)
    _summary_sheet(wb, results, len(results), has_expected, provider_name)
    _rubric_sheet(wb, provider_name, has_expected)

    os.makedirs(os.path.dirname(out_path) or ".", exist_ok=True)
    wb.save(out_path)
    return out_path


def _findings_sheet(wb, results: list[JobResult]) -> None:
    _, _, Comment, Alignment, Border, Font, PatternFill, Side = _openpyxl()
    ws = wb.create_sheet("Findings")
    headers = ["ID", "FINDING", "COMPONENT", "PROPERTY", "CHANNEL", "SEVERITY",
               "SPEC", "CODE", "DELTA", "UNIT", "PRINCIPLES", "FIX", "EVIDENCE"]
    ws.append(headers)
    _style_header(ws, len(headers))
    thin = Side(style="thin", color=RULE)
    r = 2
    for res in results:
        for f in res.findings:
            ws.cell(row=r, column=1, value=res.case_id)
            ws.cell(row=r, column=2, value=f.finding_id)
            ws.cell(row=r, column=3, value=f.node_key)
            ws.cell(row=r, column=4, value=f.prop)
            ws.cell(row=r, column=5, value=f.channel)
            ws.cell(row=r, column=6, value=f.severity)
            ws.cell(row=r, column=7, value=str(f.design_value))
            ws.cell(row=r, column=8, value=str(f.code_value))
            ws.cell(row=r, column=9, value=f.delta)
            ws.cell(row=r, column=10, value=f.delta_unit)
            ws.cell(row=r, column=11, value=", ".join(f.ux_laws))
            ws.cell(row=r, column=12, value=f.suggested_fix)
            ws.cell(row=r, column=13, value=" ".join(f.evidence))
            for col in range(1, len(headers) + 1):
                cell = ws.cell(row=r, column=col)
                cell.font = Font(name=FONT, size=9)
                cell.alignment = Alignment(vertical="top", wrap_text=col in (12, 13))
                cell.border = Border(bottom=thin)
            r += 1
    for col, width in zip("ABCDEFGHIJKLM",
                          [12, 10, 24, 22, 12, 10, 16, 16, 10, 8, 34, 48, 34]):
        ws.column_dimensions[col].width = width


def _summary_sheet(wb, results, n: int, has_expected: bool,
                   provider_name: str) -> None:
    _, _, Comment, Alignment, Border, Font, PatternFill, Side = _openpyxl()
    ws = wb.create_sheet("Summary")
    last = n + 1
    rows = [
        ("Cases evaluated", f"=COUNTA(Cases!A2:A{last})"),
        ("Audited without error", f'=COUNTIF(Cases!J2:J{last},">=0")'),
        ("", ""),
        ("Conforming (1.0)", f"=COUNTIF(Cases!J2:J{last},1)"),
        ("Partially conforming (0.5)", f"=COUNTIF(Cases!J2:J{last},0.5)"),
        ("Failing (0.0)", f"=COUNTIF(Cases!J2:J{last},0)"),
        ("Mean conformance", f"=IFERROR(AVERAGE(Cases!J2:J{last}),\"\")"),
        ("", ""),
        ("Blockers found", f"=SUM(Cases!F2:F{last})"),
        ("Major deviations", f"=SUM(Cases!G2:G{last})"),
        ("Minor deviations", f"=SUM(Cases!H2:H{last})"),
        ("Informational", f"=SUM(Cases!I2:I{last})"),
    ]
    if has_expected:
        rows += [
            ("", ""),
            ("Rows scored by a human", f'=COUNT(Cases!K2:K{last})'),
            ("Exact agreement", f"=COUNTIF(Cases!L2:L{last},1)"),
            ("Agreement rate", f'=IFERROR(COUNTIF(Cases!L2:L{last},1)/COUNT(Cases!K2:K{last}),"")'),
            ("Mean points", f'=IFERROR(AVERAGE(Cases!L2:L{last}),"")'),
        ]
    rows += [
        ("", ""),
        ("Engine", provider_name),
        ("Total runtime (s)", round(sum(r.wall_seconds for r in results), 3)),
        ("Total cost (USD)", round(sum(r.cost_usd for r in results), 6)),
    ]

    ws.append(["METRIC", "VALUE"])
    _style_header(ws, 2)
    for i, (label, value) in enumerate(rows, start=2):
        ws.cell(row=i, column=1, value=label).font = Font(
            name=FONT, size=10, bold=label and not label.startswith(" "))
        c = ws.cell(row=i, column=2, value=value)
        c.font = Font(name=FONT, size=10)
        if isinstance(value, str) and value.startswith("="):
            c.number_format = "0.00" if "AVERAGE" in value or "/" in value else "0"
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 18

    if provider_name == "offline":
        note = ws.cell(row=len(rows) + 3, column=1,
                       value="Engine 'offline' is a deterministic policy, not a language "
                             "model. Cost is a token-count proxy for comparison only.")
        note.font = Font(name=FONT, size=9, italic=True)


def _rubric_sheet(wb, provider_name: str, has_expected: bool) -> None:
    _, _, Comment, Alignment, Border, Font, PatternFill, Side = _openpyxl()
    ws = wb.create_sheet("Rubric")
    ws.append(["FIELD", "MEANING"])
    _style_header(ws, 2)
    rows = [
        ("CONFORMANCE 1.0", "No blockers and no majors. The implementation matches the "
                            "specification on every property a user would notice."),
        ("CONFORMANCE 0.5", "No blockers, but at least one major deviation: something "
                            "visible that changes hierarchy, grouping or meaning."),
        ("CONFORMANCE 0.0", "At least one blocker: a WCAG AA failure, a removed focus "
                            "state, or an interactive target under 24px."),
        ("", ""),
        ("EXPECTED", "You fill this in. The score you would give the row, on the same "
                     "0 / 0.5 / 1 scale. Leave blank if you are not grading Brush."),
        ("POINTS", "Agreement between Brush and you. 1.0 exact, 0.5 one band apart, "
                   "0.0 two bands apart."
                   if has_expected else
                   "No EXPECTED column was filled in, so this mirrors CONFORMANCE and "
                   "measures nothing about Brush's accuracy. Fill EXPECTED to score it."),
        ("", ""),
        ("Blocker", "Fails a WCAG AA success criterion, removes a state the user needs "
                    "to operate the interface, or drops a target below 24px."),
        ("Major", "Visible to an ordinary user and changes meaning or hierarchy: "
                  "colour drift ≥3.0 ΔE, spacing ≥4px on a grouping edge, type size "
                  "≥12% off, weight ≥200 off."),
        ("Minor", "Perceptible on close inspection but does not change meaning: "
                  "ΔE 1.0–3.0, spacing 2–4px, type size 6–12% off, one weight step."),
        ("Info", "Below perception but off-spec: ΔE <1.0, spacing <2px, hard-coded "
                 "literals that are not the token they sit beside."),
        ("", ""),
        ("Where these come from", "Severity bands are published in "
                                  "src/brush/knowledge/ux_laws.json and recomputed by the "
                                  "verifier from the raw measurement. Brush cannot assign "
                                  "a band the policy does not support."),
        ("Engine", f"This run used provider '{provider_name}'."
                   + (" It is a deterministic policy, not a language model."
                      if provider_name == "offline" else "")),
    ]
    for i, (a, b) in enumerate(rows, start=2):
        ws.cell(row=i, column=1, value=a).font = Font(name=FONT, size=10, bold=bool(a))
        c = ws.cell(row=i, column=2, value=b)
        c.font = Font(name=FONT, size=10)
        c.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[i].height = 30 if b else 8
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 92


def write_template(out_path: str, example_dir: str = "eval/cases") -> str:
    """A starting workbook: the legend, plus one realistic example row."""
    Workbook, _, Comment, Alignment, Border, Font, PatternFill, Side = _openpyxl()
    wb = Workbook()
    ws = wb.active
    ws.title = "Cases"
    ws.append(HEADERS)
    _style_header(ws, len(HEADERS))

    ws.append(["EXAMPLE-01",
               f"{example_dir}/design.spec.json",
               f"{example_dir}/checkout.html",
               f"{example_dir}/checkout.css",
               "", None, None, None, None,
               '=IF(F2="","",IF(F2>0,0,IF(G2>0,0.5,1)))', 1.0,
               '=IF(OR(J2="",K2=""),"",IF(ABS(J2-K2)<0.001,1,'
               'IF(ABS(ABS(J2-K2)-0.5)<0.001,0.5,0)))',
               "Example row showing the expected format. Replace or delete it."])
    for col in range(1, len(HEADERS) + 1):
        cell = ws.cell(row=2, column=col)
        cell.font = Font(name=FONT, size=10)
        cell.alignment = Alignment(vertical="top", wrap_text=col in (5, 13))
        if col in (1, 2, 3, 4, 11):
            cell.fill = PatternFill("solid", fgColor=INPUT_FILL)
    ws.row_dimensions[2].height = 40

    legend = [
        "",
        "HOW TO USE THIS SHEET",
        "Yellow cells are yours to fill in. Everything else is written by Brush.",
        "  ID              any label you want for the row",
        "  IMAGE/FIGMA     path to a design.spec.json, or to a mockup image "
        "(.png/.jpg — needs --provider anthropic to read it)",
        "  CODE            path to the .html implementation",
        "  CSS             stylesheet(s), comma separated. Left blank, Brush looks "
        "for a .css beside the HTML",
        "  EXPECTED        optional. The score you would give the row (0, 0.5 or 1). "
        "Fill it in to grade Brush; leave blank to just audit.",
        "",
        "Then run:  brush batch --sheet cases.xlsx --out results.xlsx",
        "See the Rubric tab for what the scores mean.",
    ]
    for i, line in enumerate(legend, start=4):
        c = ws.cell(row=i, column=1, value=line)
        c.font = Font(name=FONT, size=10, bold=(i == 5))
    for col, width in zip("ABCDEFGHIJKLM",
                          [14, 30, 26, 22, 30, 10, 9, 9, 8, 14, 12, 10, 62]):
        ws.column_dimensions[col].width = width

    _rubric_sheet(wb, "offline", True)
    os.makedirs(os.path.dirname(out_path) or ".", exist_ok=True)
    wb.save(out_path)
    return out_path
